from openpyxl import load_workbook


def translate_leg_name():
    wb = load_workbook('lol中英英雄.xlsx')

    sheet = wb["工作表1"]
    eng = []
    chi = []
    leg_dic = {}

    # 英文
    max_row = sheet.max_row  # 162
    # (<Cell '工作表1'.A1>,), (<Cell '工作表1'.A2>,)....
    columnA = sheet["A1":"A163"]
    for column_cells in columnA:  # (<Cell '工作表1'.A1>,) 直的
        for cell in column_cells:
            eng.append(cell.value)

    # (<Cell '工作表1'.A1>,), (<Cell '工作表1'.A2>,)....
    columnB = sheet["B1":"B163"]
    for column_cells in columnB:  # (<Cell '工作表1'.A1>,) 直的
        for cell in column_cells:
            chi.append(cell.value)

    chi.pop(17)
    chi.insert(17, "布郎姆")
    chi.pop(14)
    chi.insert(14, "貝爾薇斯")
    for i, e in enumerate(chi):
        leg_dic[e] = eng[i]
    leg_dic.update({"凱莎": "KaiSa", "努努和威朗普": "nunu", "卡力斯": "khazix", "嘉文四世": "jarvaniv",
                    "好運姐": "missfortune", "威寇茲": "velkoz", "寇格魔": "kogmaw", "悟空": "monkeyking",
                    "易大師": "masteryi", "李星": "leesin", "睿娜妲．格萊斯克": "renata", "科加斯": "chogath",
                    "翱銳龍獸": "aurelionsol", "蒙多醫生": "drmundo", "貝爾薇斯": "belveth", "貪啃奇": "tahmkench", "趙信": "xinzhao",
                    "達瑞文": "draven", "雷珂煞": "reksai"})

    return leg_dic
